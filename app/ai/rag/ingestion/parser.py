"""Backup custom parser kept for fallback reference.

Current runtime ingestion uses LangChain document loaders from
`app.ai.rag.ingestion.loaders` instead of this module.
"""

import csv
import json
from pathlib import Path

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader


def extract_text_from_file(file_path: str) -> str:
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Document file not found: {file_path}")

    extension = path.suffix.lower()
    if extension == ".pdf":
        return _extract_pdf_text(path)
    if extension == ".docx":
        return _extract_docx_text(path)
    if extension == ".txt":
        return _extract_txt_text(path)
    if extension == ".json":
        return _extract_json_text(path)
    if extension == ".csv":
        return _extract_csv_text(path)
    if extension == ".xlsx":
        return _extract_xlsx_text(path)

    raise ValueError(f"Unsupported file extension for parsing: {extension}")


def _extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    pages: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            pages.append(text.strip())
    return "\n\n".join(pages)


def _extract_docx_text(path: Path) -> str:
    document = DocxDocument(str(path))
    paragraphs = [
        paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()
    ]
    return "\n\n".join(paragraphs)


def _extract_txt_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="latin-1")


def _extract_json_text(path: Path) -> str:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    return json.dumps(payload, indent=2, ensure_ascii=False)


def _extract_csv_text(path: Path) -> str:
    lines: list[str] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            lines.append(", ".join(str(cell) for cell in row))
    return "\n".join(lines)


def _extract_xlsx_text(path: Path) -> str:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sections: list[str] = []

    for sheet in workbook.worksheets:
        row_lines: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cell_values = [str(value) for value in row if value is not None and str(value).strip()]
            if cell_values:
                row_lines.append(" | ".join(cell_values))
        if row_lines:
            sections.append(f"[Sheet: {sheet.title}]\n" + "\n".join(row_lines))

    workbook.close()
    return "\n\n".join(sections)
